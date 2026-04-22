#!/usr/bin/env python3
"""
Build hotlines.xlsx from hotlines.json.

Sheets:
  1. Hotlines — one row per (country, hotline, number) with filters and a frozen header row.
  2. By Country — one row per country summarising which categories are covered.
  3. Categories — legend/reference for the category enum.

Usage (run under WSL):
    pip install openpyxl --break-system-packages
    python3 scripts/build_xlsx.py
"""
from __future__ import annotations

import json
import pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = pathlib.Path(__file__).parent.parent
SRC = ROOT / "hotlines.json"
OUT = ROOT / "hotlines.xlsx"


def main():
    data = json.loads(SRC.read_text(encoding="utf-8"))
    wb = Workbook()

    # -------- Sheet 1: Hotlines --------
    ws = wb.active
    ws.title = "Hotlines"
    header = [
        "Country", "Alpha-2", "Alpha-3", "Region", "Subregion",
        "Hotline", "Category", "Number", "Number Type",
        "Chat URL", "Email", "Website",
        "Hours", "Languages", "Cost", "Target", "Geography",
        "Notes", "Verification Status", "Last Verified", "Sources",
    ]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    row = 2
    for country in data["countries"]:
        for h in country.get("hotlines", []):
            entries = []
            for n in h.get("voice_numbers", []): entries.append(("voice", n))
            for n in h.get("sms_numbers", []): entries.append(("sms", n))
            for n in h.get("text_numbers", []): entries.append(("text", n))
            for n in h.get("short_codes", []): entries.append(("short_code", n))
            if not entries:
                entries = [("", "")]
            for ntype, number in entries:
                ws.cell(row=row, column=1, value=country["country"])
                ws.cell(row=row, column=2, value=country.get("alpha-2"))
                ws.cell(row=row, column=3, value=country.get("alpha-3"))
                ws.cell(row=row, column=4, value=country.get("region"))
                ws.cell(row=row, column=5, value=country.get("subregion"))
                ws.cell(row=row, column=6, value=h.get("name"))
                ws.cell(row=row, column=7, value=h.get("category"))
                ws.cell(row=row, column=8, value=number)
                ws.cell(row=row, column=9, value=ntype)
                ws.cell(row=row, column=10, value=h.get("chat_url"))
                ws.cell(row=row, column=11, value=h.get("email"))
                ws.cell(row=row, column=12, value=h.get("website"))
                ws.cell(row=row, column=13, value=h.get("hours"))
                ws.cell(row=row, column=14, value=", ".join(h.get("languages", []) or []))
                ws.cell(row=row, column=15, value=h.get("cost"))
                ws.cell(row=row, column=16, value=h.get("target"))
                ws.cell(row=row, column=17, value=h.get("geography"))
                ws.cell(row=row, column=18, value=h.get("notes"))
                ws.cell(row=row, column=19, value=h.get("verification_status"))
                ws.cell(row=row, column=20, value=h.get("last_verified"))
                ws.cell(row=row, column=21, value=", ".join(h.get("sources", []) or []))
                row += 1

    ws.freeze_panes = "A2"
    widths = {1: 24, 2: 8, 3: 8, 4: 14, 5: 20, 6: 36, 7: 18, 8: 20, 9: 12, 10: 32, 11: 24, 12: 32, 13: 20, 14: 22, 15: 14, 16: 34, 17: 20, 18: 40, 19: 18, 20: 14, 21: 40}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    last_col = get_column_letter(len(header))
    table_ref = f"A1:{last_col}{row-1}"
    tbl = Table(displayName="Hotlines", ref=table_ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)

    # -------- Sheet 2: By Country --------
    ws2 = wb.create_sheet("By Country")
    categories = sorted(data.get("categories_reference", {}).keys())
    header2 = ["Country", "Alpha-2", "Region", "General Emergency", "Hotline Count"] + categories
    ws2.append(header2)
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for country in sorted(data["countries"], key=lambda c: c["country"]):
        cats = {}
        for h in country.get("hotlines", []):
            cats[h.get("category")] = cats.get(h.get("category"), 0) + 1
        r = [
            country["country"], country.get("alpha-2"), country.get("region"),
            ", ".join(country.get("general_emergency", [])),
            len(country.get("hotlines", [])),
        ]
        for cat in categories:
            r.append(cats.get(cat, 0) or "")
        ws2.append(r)
    ws2.freeze_panes = "B2"
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 8
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 20
    ws2.column_dimensions["E"].width = 14
    for i in range(len(categories)):
        ws2.column_dimensions[get_column_letter(6 + i)].width = 16

    # -------- Sheet 3: Categories --------
    ws3 = wb.create_sheet("Categories")
    ws3.append(["Category", "Description"])
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
    for k, v in sorted(data.get("categories_reference", {}).items()):
        ws3.append([k, v])
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 70

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
